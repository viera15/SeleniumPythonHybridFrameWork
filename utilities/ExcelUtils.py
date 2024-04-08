import openpyxl

def get_row_count(path, sheet_name):
    '''Zistí maximálny počet vyplnených riadkov'''
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(path, sheet_name):
    '''Zístí maximálny počet vyplnených stĺpcov'''
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def get_cell_data(path, sheet_name, row_number, column_number):
    '''Zistí hodnotu bunky(obsah)'''
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=column_number).value


def set_cell_data(path, sheet_name, row_number, column_number, data):
    '''Vloží dáta do bunky'''
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(path)

def get_data_from_excel(path, sheet_name):
    '''Funkcia načíta dáta z Excelu do listu/listov'''
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    row_count = sheet.max_row
    col_count = sheet.max_column
    for r in range(2, row_count+1):
        row_list = []
        for c in range(1, col_count+1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)
    return final_list

